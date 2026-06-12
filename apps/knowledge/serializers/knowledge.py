import io
import json
import os
import re
import zipfile
from collections import defaultdict
from functools import reduce
from tempfile import TemporaryDirectory
from typing import Dict, List
from urllib.parse import quote

import uuid_utils.compat as uuid
from celery_once import AlreadyQueued
from common.chunk import text_to_chunk
from common.config.embedding_config import VectorStore
from common.database_model_manage.database_model_manage import DatabaseModelManage
from common.db.search import get_dynamics_model, native_page_search, native_search
from common.db.sql_execute import select_list
from common.event.listener_manage import ListenerManagement
from common.exception.app_exception import AppApiException
from common.field.common import UploadedFileField
from common.utils.common import bulk_create_in_batches, get_file_content, parse_image, post
from django.core import validators
from django.db import models, transaction
from django.db.models import QuerySet
from django.db.models.functions import Reverse, Substr
from django.db.models.query_utils import Q
from django.http import HttpResponse
from django.utils.translation import gettext
from django.utils.translation import gettext_lazy as _
from maxkb.conf import PROJECT_DIR
from models_provider.models import Model
from rest_framework import serializers
from system_manage.models import AuthTargetType, WorkspaceUserResourcePermission
from system_manage.models.resource_mapping import ResourceMapping
from system_manage.serializers.resource_mapping_serializers import ResourceMappingSerializer
from system_manage.serializers.user_resource_permission import UserResourcePermissionSerializer
from users.serializers.user import is_workspace_manage, is_workspace_manage_permission_read

from knowledge.models import (
    Document,
    DocumentTag,
    File,
    FileSourceType,
    Knowledge,
    KnowledgeFolder,
    KnowledgeScope,
    KnowledgeType,
    Paragraph,
    Problem,
    ProblemParagraphMapping,
    SearchMode,
    State,
    Tag,
    TaskType,
    Termbase,
)
from knowledge.serializers.common import (
    BatchMoveSerializer,
    BatchSerializer,
    GenerateRelatedSerializer,
    MetaSerializer,
    ProblemParagraphManage,
    ProblemParagraphObject,
    drop_knowledge_index,
    get_embedding_model_by_knowledge_id,
    get_embedding_model_by_knowledge_id_list,
    get_embedding_model_id_by_knowledge_id,
    list_paragraph,
    update_resource_mapping_by_knowledge,
    write_image,
    zip_dir,
)
from knowledge.serializers.document import DocumentSerializers
from knowledge.task.embedding import delete_embedding_by_knowledge, embedding_by_knowledge
from knowledge.task.generate import generate_related_by_knowledge_id


class KnowledgeModelSerializer(serializers.ModelSerializer):
    class Meta:
        model = Knowledge
        fields = [
            "id",
            "name",
            "desc",
            "meta",
            "folder_id",
            "type",
            "workspace_id",
            "create_time",
            "update_time",
            "file_size_limit",
            "file_count_limit",
            "embedding_model_id",
        ]


class KnowledgeBaseCreateRequest(serializers.Serializer):
    name = serializers.CharField(required=True, label=_("knowledge name"))
    folder_id = serializers.CharField(required=True, label=_("folder id"))
    desc = serializers.CharField(required=False, allow_null=True, allow_blank=True, label=_("knowledge description"))
    embedding_model_id = serializers.CharField(required=True, label=_("knowledge embedding"))


class KnowledgeImportRequest(serializers.Serializer):
    file = UploadedFileField(required=True, label=_("file"))


class KnowledgeEditRequest(serializers.Serializer):
    name = serializers.CharField(required=False, max_length=64, min_length=1, label=_("knowledge name"))
    desc = serializers.CharField(required=False, max_length=256, min_length=1, label=_("knowledge description"))
    meta = serializers.DictField(required=False)
    application_id_list = serializers.ListSerializer(
        required=False,
        child=serializers.UUIDField(required=True, label=_("application id")),
        label=_("application id list"),
    )
    file_size_limit = serializers.IntegerField(required=False, label=_("file size limit"))
    file_count_limit = serializers.IntegerField(required=False, label=_("file count limit"))

    @staticmethod
    def get_knowledge_meta_valid_map():
        knowledge_meta_valid_map = {
            KnowledgeType.BASE: MetaSerializer.BaseMeta,
        }
        return knowledge_meta_valid_map

    def is_valid(self, *, knowledge: Knowledge = None):
        super().is_valid(raise_exception=True)
        if "meta" in self.data and self.data.get("meta") is not None:
            knowledge_meta_valid_map = self.get_knowledge_meta_valid_map()
            valid_class = knowledge_meta_valid_map.get(knowledge.type)
            valid_class(data=self.data.get("meta")).is_valid(raise_exception=True)


class HitTestSerializer(serializers.Serializer):
    query_text = serializers.CharField(required=True, label=_("query text"))
    top_number = serializers.IntegerField(required=True, max_value=10000, min_value=1, label=_("top number"))
    similarity = serializers.FloatField(required=True, max_value=2, min_value=0, label=_("similarity"))
    search_mode = serializers.CharField(
        required=True,
        label=_("search mode"),
        validators=[
            validators.RegexValidator(
                regex=re.compile("^embedding|keywords|blend$"),
                message=_("The type only supports embedding|keywords|blend"),
                code=500,
            )
        ],
    )


class BatchHitTestSerializer(HitTestSerializer):
    knowledge_id_list = serializers.ListField(
        required=True,
        allow_empty=False,
        child=serializers.UUIDField(required=True, label=_("knowledge id")),
        label=_("knowledge id list"),
    )


class KnowledgeSerializer(serializers.Serializer):
    class Query(serializers.Serializer):
        workspace_id = serializers.CharField(required=True)
        folder_id = serializers.CharField(required=False, label=_("folder id"), allow_null=True)
        name = serializers.CharField(
            required=False, label=_("knowledge name"), allow_null=True, allow_blank=True, max_length=64, min_length=1
        )
        desc = serializers.CharField(
            required=False,
            label=_("knowledge description"),
            allow_null=True,
            allow_blank=True,
            max_length=256,
            min_length=1,
        )
        user_id = serializers.UUIDField(required=False, label=_("user id"), allow_null=True)
        scope = serializers.CharField(required=False, label=_("knowledge scope"), allow_null=True)
        create_user = serializers.UUIDField(required=False, label=_("create user"), allow_null=True)

        @staticmethod
        def is_x_pack_ee():
            workspace_user_role_mapping_model = DatabaseModelManage.get_model("workspace_user_role_mapping")
            role_permission_mapping_model = DatabaseModelManage.get_model("role_permission_mapping_model")
            return workspace_user_role_mapping_model is not None and role_permission_mapping_model is not None

        def get_query_set(self, workspace_manage, is_x_pack_ee):
            self.is_valid(raise_exception=True)
            workspace_id = self.data.get("workspace_id")
            query_set_dict = {}
            query_set = QuerySet(
                model=get_dynamics_model(
                    {
                        "temp.name": models.CharField(),
                        "temp.desc": models.CharField(),
                        "document_temp.char_length": models.IntegerField(),
                        "temp.create_time": models.DateTimeField(),
                        "temp.user_id": models.CharField(),
                        "temp.workspace_id": models.CharField(),
                        "temp.folder_id": models.CharField(),
                        "temp.id": models.CharField(),
                        "temp.scope": models.CharField(),
                    }
                )
            )
            folder_query_set = QuerySet(KnowledgeFolder)

            if "desc" in self.data and self.data.get("desc") is not None:
                query_set = query_set.filter(**{"temp.desc__icontains": self.data.get("desc")})
                folder_query_set = folder_query_set.filter(**{"desc__icontains": self.data.get("desc")})
            if "name" in self.data and self.data.get("name") is not None:
                query_set = query_set.filter(**{"temp.name__icontains": self.data.get("name")})
                folder_query_set = folder_query_set.filter(**{"name__icontains": self.data.get("name")})
            if "workspace_id" in self.data and self.data.get("workspace_id") is not None:
                query_set = query_set.filter(**{"temp.workspace_id": self.data.get("workspace_id")})
                folder_query_set = folder_query_set.filter(**{"workspace_id": self.data.get("workspace_id")})
            if (
                "folder_id" in self.data
                and self.data.get("folder_id") is not None
                and self.data.get("workspace_id") != self.data.get("folder_id")
            ):
                query_set = query_set.filter(**{"temp.folder_id": self.data.get("folder_id")})
                folder_query_set = folder_query_set.filter(**{"parent_id": self.data.get("folder_id")})
            if "scope" in self.data and self.data.get("scope") is not None:
                query_set = query_set.filter(**{"temp.scope": self.data.get("scope")})
            if "create_user" in self.data and self.data.get("create_user") is not None:
                query_set = query_set.filter(**{"temp.user_id": self.data.get("create_user")})
            query_set = query_set.order_by("-temp.create_time", "temp.id")
            query_set_dict["default_sql"] = query_set

            query_set_dict["knowledge_custom_sql"] = QuerySet(
                model=get_dynamics_model(
                    {
                        "knowledge.workspace_id": models.CharField(),
                    }
                )
            ).filter(**{"knowledge.workspace_id": workspace_id})
            # query_set_dict['folder_query_set'] = folder_query_set
            if not workspace_manage:
                query_set_dict["workspace_user_resource_permission_query_set"] = QuerySet(
                    WorkspaceUserResourcePermission
                ).filter(auth_target_type="KNOWLEDGE", workspace_id=workspace_id, user_id=self.data.get("user_id"))
            return query_set_dict

        def page(self, current_page: int, page_size: int):
            self.is_valid(raise_exception=True)
            folder_id = self.data.get("folder_id", self.data.get("workspace_id"))
            root = KnowledgeFolder.objects.filter(id=folder_id).first()
            if not root:
                raise serializers.ValidationError(_("Folder not found"))
            workspace_manage = is_workspace_manage_permission_read(
                self.data.get("user_id"), self.data.get("workspace_id"), "KNOWLEDGE:READ"
            )
            is_x_pack_ee = self.is_x_pack_ee()
            result = native_page_search(
                current_page,
                page_size,
                self.get_query_set(workspace_manage, is_x_pack_ee),
                select_string=get_file_content(
                    os.path.join(
                        PROJECT_DIR,
                        "apps",
                        "knowledge",
                        "sql",
                        "list_knowledge.sql"
                        if workspace_manage
                        else ("list_knowledge_user_ee.sql" if is_x_pack_ee else "list_knowledge_user.sql"),
                    )
                ),
                post_records_handler=lambda r: r,
            )
            return ResourceMappingSerializer().get_resource_count(result)

        def list(self):
            self.is_valid(raise_exception=True)
            folder_id = self.data.get("folder_id")
            if not folder_id:
                folder_id = self.data.get("workspace_id")
            root = KnowledgeFolder.objects.filter(id=folder_id).first()
            if not root:
                raise serializers.ValidationError(_("Folder not found"))
            workspace_manage = is_workspace_manage_permission_read(
                self.data.get("user_id"), self.data.get("workspace_id"), "KNOWLEDGE:READ"
            )

            is_x_pack_ee = self.is_x_pack_ee()
            return native_search(
                self.get_query_set(workspace_manage, is_x_pack_ee),
                select_string=get_file_content(
                    os.path.join(
                        PROJECT_DIR,
                        "apps",
                        "knowledge",
                        "sql",
                        "list_knowledge.sql"
                        if workspace_manage
                        else ("list_knowledge_user_ee.sql" if self.is_x_pack_ee() else "list_knowledge_user.sql"),
                    )
                ),
            )

    class Operate(serializers.Serializer):
        user_id = serializers.UUIDField(required=True, label=_("user id"))
        workspace_id = serializers.CharField(required=True, label=_("workspace id"))
        knowledge_id = serializers.UUIDField(required=True, label=_("knowledge id"))

        @staticmethod
        def _parse_boolean_param(value, field_name):
            if value in serializers.BooleanField.TRUE_VALUES:
                return True
            if value in serializers.BooleanField.FALSE_VALUES or value is None:
                return False
            raise AppApiException(500, _("%s must be a boolean") % field_name)

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            workspace_id = self.data.get("workspace_id")
            query_set = QuerySet(Knowledge).filter(id=self.data.get("knowledge_id"))
            if workspace_id:
                query_set = query_set.filter(workspace_id=workspace_id)
            if not query_set.exists():
                raise AppApiException(500, _("Knowledge id does not exist"))

        @transaction.atomic
        def embedding(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            knowledge_id = self.data.get("knowledge_id")
            knowledge = QuerySet(Knowledge).filter(id=knowledge_id).first()
            embedding_model_id = knowledge.embedding_model_id
            embedding_model = QuerySet(Model).filter(id=embedding_model_id).first()
            if embedding_model is None:
                raise AppApiException(500, _("Model does not exist"))
            ListenerManagement.update_status(
                QuerySet(Document).filter(knowledge_id=self.data.get("knowledge_id")), TaskType.EMBEDDING, State.PENDING
            )
            ListenerManagement.update_status(
                QuerySet(Paragraph).filter(knowledge_id=self.data.get("knowledge_id")),
                TaskType.EMBEDDING,
                State.PENDING,
            )
            ListenerManagement.get_aggregation_document_status_by_knowledge_id(self.data.get("knowledge_id"))()
            embedding_model_id = get_embedding_model_id_by_knowledge_id(self.data.get("knowledge_id"))
            try:
                embedding_by_knowledge.delay(knowledge_id, embedding_model_id)
            except AlreadyQueued:
                raise AppApiException(500, _("Failed to send the vectorization task, please try again later!"))

        def generate_related(self, instance: Dict, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
                GenerateRelatedSerializer(data=instance).is_valid(raise_exception=True)
            knowledge_id = self.data.get("knowledge_id")
            model_id = instance.get("model_id")
            prompt = instance.get("prompt")
            model_params_setting = instance.get("model_params_setting")
            state_list = instance.get("state_list")
            ListenerManagement.update_status(
                QuerySet(Document).filter(knowledge_id=knowledge_id), TaskType.GENERATE_PROBLEM, State.PENDING
            )
            ListenerManagement.update_status(
                QuerySet(Paragraph)
                .annotate(
                    reversed_status=Reverse("status"),
                    task_type_status=Substr("reversed_status", TaskType.GENERATE_PROBLEM.value, 1),
                )
                .filter(task_type_status__in=state_list, knowledge_id=knowledge_id)
                .values("id"),
                TaskType.GENERATE_PROBLEM,
                State.PENDING,
            )
            ListenerManagement.get_aggregation_document_status_by_knowledge_id(knowledge_id)()
            try:
                generate_related_by_knowledge_id.delay(knowledge_id, model_id, model_params_setting, prompt, state_list)
            except AlreadyQueued:
                raise AppApiException(500, _("Failed to send the vectorization task, please try again later!"))

        def list_application(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            # knowledge = QuerySet(Knowledge).get(id=self.data.get("knowledge_id"))
            return select_list(
                get_file_content(
                    os.path.join(PROJECT_DIR, "apps", "knowledge", "sql", "list_knowledge_application.sql")
                ),
                [
                    self.data.get("user_id"),
                ],
            )

        @staticmethod
        def is_x_pack_ee():
            workspace_user_role_mapping_model = DatabaseModelManage.get_model("workspace_user_role_mapping")
            role_permission_mapping_model = DatabaseModelManage.get_model("role_permission_mapping_model")
            return workspace_user_role_mapping_model is not None and role_permission_mapping_model is not None

        def one(self):
            self.is_valid()
            workspace_manage = is_workspace_manage(self.data.get("user_id"), self.data.get("workspace_id"))
            is_x_pack_ee = self.is_x_pack_ee()

            query_set_dict = {
                "default_sql": QuerySet(model=get_dynamics_model({"temp.id": models.CharField()})).filter(
                    **{"temp.id": self.data.get("knowledge_id")}
                ),
                "knowledge_custom_sql": QuerySet(model=get_dynamics_model({"knowledge.id": models.CharField()})).filter(
                    **{"knowledge.id": self.data.get("knowledge_id")}
                ),
            }
            if not workspace_manage:
                query_set_dict["workspace_user_resource_permission_query_set"] = QuerySet(
                    WorkspaceUserResourcePermission
                ).filter(
                    auth_target_type="KNOWLEDGE",
                    workspace_id=self.data.get("workspace_id"),
                    user_id=self.data.get("user_id"),
                )
            all_application_list = [str(adm.get("id")) for adm in self.list_application(with_valid=False)]
            knowledge_dict = native_search(
                query_set_dict,
                select_string=get_file_content(
                    os.path.join(
                        PROJECT_DIR,
                        "apps",
                        "knowledge",
                        "sql",
                        "list_knowledge.sql"
                        if workspace_manage
                        else ("list_knowledge_user_ee.sql" if is_x_pack_ee else "list_knowledge_user.sql"),
                    )
                ),
                with_search_one=True,
            )
            return {
                **knowledge_dict,
                "meta": json.loads(knowledge_dict.get("meta", "{}")),
                "application_id_list": list(
                    filter(
                        lambda application_id: all_application_list.__contains__(application_id),
                        [
                            str(application_knowledge_mapping.source_id)
                            for application_knowledge_mapping in QuerySet(ResourceMapping).filter(
                                source_type="APPLICATION",
                                target_type="KNOWLEDGE",
                                target_id=self.data.get("knowledge_id"),
                            )
                        ],
                    )
                ),
            }

        @transaction.atomic
        def edit(self, instance: Dict, select_one=True):
            self.is_valid()
            knowledge = QuerySet(Knowledge).get(id=self.data.get("knowledge_id"))
            KnowledgeEditRequest(data=instance).is_valid(knowledge=knowledge)
            if "embedding_model_id" in instance:
                knowledge.embedding_model_id = instance.get("embedding_model_id")
            if "name" in instance:
                knowledge.name = instance.get("name")
            if "desc" in instance:
                knowledge.desc = instance.get("desc")
            if "meta" in instance:
                knowledge.meta = instance.get("meta")
            if "folder_id" in instance:
                knowledge.folder_id = instance.get("folder_id")
            if "file_size_limit" in instance:
                knowledge.file_size_limit = instance.get("file_size_limit")
            if "file_count_limit" in instance:
                knowledge.file_count_limit = instance.get("file_count_limit")
            knowledge.save()
            update_resource_mapping_by_knowledge(str(knowledge.id))
            if select_one:
                return self.one()
            return None

        @transaction.atomic
        def delete(self):
            self.is_valid()
            knowledge = QuerySet(Knowledge).get(id=self.data.get("knowledge_id"))
            QuerySet(Document).filter(knowledge=knowledge).delete()
            QuerySet(ProblemParagraphMapping).filter(knowledge=knowledge).delete()
            QuerySet(Paragraph).filter(knowledge=knowledge).delete()
            QuerySet(Problem).filter(knowledge=knowledge).delete()
            QuerySet(WorkspaceUserResourcePermission).filter(target=knowledge.id).delete()
            drop_knowledge_index(knowledge_id=knowledge.id)
            knowledge.delete()
            File.objects.filter(
                source_id=knowledge.id,
            ).delete()
            QuerySet(ResourceMapping).filter(
                Q(target_id=self.data.get("knowledge_id")) | Q(source_id=self.data.get("knowledge_id"))
            ).delete()
            delete_embedding_by_knowledge(self.data.get("knowledge_id"))
            return True

        def export_excel(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            document_list = QuerySet(Document).filter(knowledge_id=self.data.get("knowledge_id"))
            paragraph_list = native_search(
                QuerySet(Paragraph).filter(knowledge_id=self.data.get("knowledge_id")),
                get_file_content(
                    os.path.join(PROJECT_DIR, "apps", "knowledge", "sql", "list_paragraph_document_name.sql")
                ),
            )
            problem_mapping_list = native_search(
                QuerySet(ProblemParagraphMapping).filter(knowledge_id=self.data.get("knowledge_id")),
                get_file_content(os.path.join(PROJECT_DIR, "apps", "knowledge", "sql", "list_problem_mapping.sql")),
                with_table_name=True,
            )
            data_dict, document_dict = DocumentSerializers.Operate.merge_problem(
                paragraph_list, problem_mapping_list, document_list
            )
            workbook = DocumentSerializers.Operate.get_workbook(data_dict, document_dict)
            response = HttpResponse(content_type="application/vnd.ms-excel")
            response["Content-Disposition"] = 'attachment; filename="knowledge.xlsx"'
            workbook.save(response)
            return response

        def export_zip(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            knowledge = QuerySet(Knowledge).filter(id=self.data.get("knowledge_id")).first()
            document_list = QuerySet(Document).filter(knowledge_id=self.data.get("knowledge_id"))
            paragraph_list = native_search(
                QuerySet(Paragraph).filter(knowledge_id=self.data.get("knowledge_id")),
                get_file_content(
                    os.path.join(PROJECT_DIR, "apps", "knowledge", "sql", "list_paragraph_document_name.sql")
                ),
            )
            problem_mapping_list = native_search(
                QuerySet(ProblemParagraphMapping).filter(knowledge_id=self.data.get("knowledge_id")),
                get_file_content(os.path.join(PROJECT_DIR, "apps", "knowledge", "sql", "list_problem_mapping.sql")),
                with_table_name=True,
            )
            data_dict, document_dict = DocumentSerializers.Operate.merge_problem(
                paragraph_list, problem_mapping_list, document_list
            )
            res = [parse_image(paragraph.get("content")) for paragraph in paragraph_list]

            workbook = DocumentSerializers.Operate.get_workbook(data_dict, document_dict)
            response = HttpResponse(content_type="application/zip")
            response["Content-Disposition"] = f'attachment; filename="{knowledge.name}.zip"'
            zip_buffer = io.BytesIO()
            with TemporaryDirectory() as tempdir:
                knowledge_file = os.path.join(tempdir, "knowledge.xlsx")
                workbook.save(knowledge_file)
                for r in res:
                    write_image(tempdir, r)
                zip_dir(tempdir, zip_buffer)
            response.write(zip_buffer.getvalue())
            return response

        def export_knowledge(self, with_source_file=False, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            with_source_file = self._parse_boolean_param(with_source_file, "with_source_file")
            knowledge_id = self.data.get("knowledge_id")
            knowledge = QuerySet(Knowledge).filter(id=knowledge_id).first()

            document_list = QuerySet(Document).filter(knowledge_id=knowledge_id)
            paragraph_list = native_search(
                QuerySet(Paragraph).filter(knowledge_id=self.data.get("knowledge_id")).order_by("position"),
                get_file_content(
                    os.path.join(PROJECT_DIR, "apps", "knowledge", "sql", "list_paragraph_document_name.sql")
                ),
            )
            problem_mapping_list = native_search(
                QuerySet(ProblemParagraphMapping).filter(knowledge_id=self.data.get("knowledge_id")),
                get_file_content(os.path.join(PROJECT_DIR, "apps", "knowledge", "sql", "list_problem_mapping.sql")),
                with_table_name=True,
            )
            data_dict, document_dict = DocumentSerializers.Operate.merge_problem(
                paragraph_list, problem_mapping_list, document_list
            )
            source_file_list = []
            if with_source_file:
                document_id_list = [str(document.id) for document in document_list]
                source_file_list = list(
                    QuerySet(File).filter(source_id__in=document_id_list, source_type=FileSourceType.DOCUMENT)
                )
            source_file_map = {str(source_file.source_id): source_file for source_file in source_file_list}

            # 查询标签和文档标签关联
            tag_list = list(QuerySet(Tag).filter(knowledge_id=knowledge_id).values("id", "key", "value"))
            document_tag_list = list(
                QuerySet(DocumentTag).filter(document__knowledge_id=knowledge_id).values("document_id", "tag_id")
            )
            # 知识库标签map
            tag_map = {t["id"]: t for t in tag_list}
            # 文档标签map
            doc_tag_map = defaultdict(list)

            for dt in document_tag_list:
                tag = tag_map.get(dt["tag_id"])
                if tag:
                    doc_tag_map[dt["document_id"]].append(f"{tag['key']}:{tag['value']}")

            # doc_id -> document_obj
            doc_obj_map = {}
            for doc in document_list:
                if with_source_file:
                    doc.meta = {**doc.meta} if doc.meta else {}
                    source_file = source_file_map.get(str(doc.id))
                    if source_file:
                        doc.meta["source_file_id"] = str(source_file.id)
                    else:
                        doc.meta.pop("source_file_id", None)
                doc_obj_map[doc.id] = doc

            # termbase
            terms = list(
                QuerySet(Termbase)
                .filter(
                    knowledge_id=knowledge_id,
                )
                .values_list("content", flat=True)
            )

            # paragraph_id -> is_active
            paragraph_active_map = {}
            for p in paragraph_list:
                doc_id = p.get("document_id")
                if doc_id not in paragraph_active_map:
                    paragraph_active_map[doc_id] = []
                paragraph_active_map[doc_id].append("1" if p.get("is_active") else "0")

            res = [parse_image(paragraph.get("content")) for paragraph in paragraph_list]
            # 新增字段
            workbook = self._get_knowledge_workbook(
                data_dict, document_dict, doc_tag_map, doc_obj_map, paragraph_active_map
            )

            response = HttpResponse(content_type="application/zip")
            response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(knowledge.name)}.zip"
            zip_buffer = io.BytesIO()
            with TemporaryDirectory() as tempdir:
                knowledge_file_path = os.path.join(tempdir, "knowledge.xlsx")
                workbook.save(knowledge_file_path)

                for r in res:
                    write_image(tempdir, r)

                source_file_path_set = set()
                source_file_export_list = []
                document_sheet_name_map = {
                    str(document_id): sheet_name for document_id, sheet_name in document_dict.items()
                }
                for source_file in source_file_list:
                    source_file_zip_path = self._get_source_file_zip_path(source_file.file_name, source_file_path_set)
                    source_file_export_list.append(
                        {
                            "id": str(source_file.id),
                            "file_name": source_file.file_name,
                            "source_id": source_file.source_id,
                            "sheet_name": document_sheet_name_map.get(str(source_file.source_id)),
                            "zip_path": source_file_zip_path,
                        }
                    )
                    source_file_path = os.path.join(tempdir, source_file_zip_path)
                    os.makedirs(os.path.dirname(source_file_path), exist_ok=True)
                    with open(source_file_path, "wb") as f:
                        f.write(source_file.get_bytes())

                knowledge_json = {
                    "name": knowledge.name,
                    "desc": knowledge.desc,
                    "type": knowledge.type,
                    "meta": {} if knowledge.type == KnowledgeType.LARK else (knowledge.meta if knowledge.meta else {}),
                    "file_size_limit": knowledge.file_size_limit,
                    "file_count_limit": knowledge.file_count_limit,
                    "tags": [{"key": t["key"], "value": t["value"]} for t in tag_list],
                    "termbase": terms,
                    "source_file_list": source_file_export_list,
                }

                with open(os.path.join(tempdir, "knowledge.json"), "w", encoding="utf-8") as f:
                    json.dump(knowledge_json, f, ensure_ascii=False)

                zip_dir(tempdir, zip_buffer)
            response.write(zip_buffer.getvalue())
            return response

        @staticmethod
        def _get_knowledge_workbook(
            data_dict: dict, document_dict: dict, doc_tag_map: dict, doc_obj_map: dict, paragraph_active_map: dict
        ):
            import openpyxl
            from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
            if len(data_dict.keys()) == 0:
                data_dict["sheet"] = []
            for sheet_id in data_dict:
                sheet_name = document_dict.get(sheet_id)
                worksheet = workbook.create_sheet(sheet_name)

                doc = doc_obj_map.get(sheet_id) if sheet_id in doc_obj_map else None
                tags_str = "|".join(doc_tag_map.get(sheet_id, []))
                hit_method = doc.hit_handling_method if doc else ""
                similarity = doc.directly_return_similarity if doc else ""
                is_active = "1" if (doc and doc.is_active) else "0"
                doc_type = doc.type if doc else ""
                doc_meta = json.dumps(doc.meta, ensure_ascii=False) if (doc and doc.meta) else ""

                header = [
                    gettext("Section title (optional)"),
                    gettext("Section content (required, question answer, no more than 4096 characters)"),
                    gettext("Question (optional, one per line in the cell)"),
                    gettext("Tags"),
                    gettext("Hit handling method"),
                    gettext("Directly return similarity"),
                    gettext("Is active"),
                    gettext("Paragraph is active"),
                    gettext("Document type"),
                    gettext("Document meta"),
                ]

                rows = data_dict.get(sheet_id, [])
                para_active_list = paragraph_active_map.get(sheet_id, [])
                # 初始化标题
                data = [header]
                for row_idx, row in enumerate(rows):
                    para_active = para_active_list[row_idx] if row_idx < len(para_active_list) else "1"
                    # None 转为 ''
                    row = [col if col is not None else "" for col in row]
                    # 补齐到3列
                    row = (row + ["", "", ""])[:3]
                    if row_idx == 0:
                        data.append(
                            [*row, tags_str, hit_method, similarity, is_active, para_active, doc_type, doc_meta]
                        )
                    else:
                        data.append([*row, "", "", "", "", para_active, "", ""])

                for row_idx, row in enumerate(data):
                    for col_idx, col in enumerate(row):
                        cell = worksheet.cell(row=row_idx + 1, column=col_idx + 1)
                        if isinstance(col, str):
                            col = re.sub(ILLEGAL_CHARACTERS_RE, "", col)
                            if col.startswith(("=", "+", "-", "@")):
                                col = "\ufeff" + col
                        cell.value = col
            return workbook

        @staticmethod
        def _get_source_file_zip_path(file_name, source_file_path_set):
            file_name = file_name.replace("\\", "/") if file_name else ""
            file_name = os.path.basename(file_name).strip() or "source_file"
            name, ext = os.path.splitext(file_name)
            source_file_path = os.path.join("source_file", file_name)
            index = 1
            while source_file_path in source_file_path_set:
                source_file_path = os.path.join("source_file", f"{name}({index}){ext}")
                index += 1
            source_file_path_set.add(source_file_path)
            return source_file_path

        @staticmethod
        def _restore_source_file(zf, namelist_set, source_file_id, source_file_meta, document_id):
            source_file_bytes = None
            source_file_path_list = []
            if source_file_meta and source_file_meta.get("zip_path"):
                source_file_path_list.append(source_file_meta.get("zip_path"))
            source_file_path_list.append(os.path.join("source_file", source_file_id))
            for source_file_path in source_file_path_list:
                if source_file_path in namelist_set:
                    source_file_bytes = zf.read(source_file_path)
                    break
            else:
                old_file = QuerySet(File).filter(id=source_file_id).first()
                if old_file:
                    source_file_bytes = old_file.get_bytes()
                    if source_file_meta is None:
                        source_file_meta = {"file_name": old_file.file_name}
            if source_file_bytes is None:
                return None

            source_file = File(
                id=uuid.uuid7(),
                file_name=(
                    source_file_meta.get("file_name")
                    if source_file_meta and source_file_meta.get("file_name")
                    else source_file_id
                ),
                source_type=FileSourceType.DOCUMENT,
                source_id=document_id,
                meta={},
            )
            source_file.save(source_file_bytes)
            return source_file.id

        @staticmethod
        def merge_problem(paragraph_list: List[Dict], problem_mapping_list: List[Dict]):
            result = {}
            document_dict = {}

            for paragraph in paragraph_list:
                problem_list = [
                    problem_mapping.get("content")
                    for problem_mapping in problem_mapping_list
                    if problem_mapping.get("paragraph_id") == paragraph.get("id")
                ]
                document_sheet = result.get(paragraph.get("document_id"))
                d = document_dict.get(paragraph.get("document_name"))
                if d is None:
                    document_dict[paragraph.get("document_name")] = {paragraph.get("document_id")}
                else:
                    d.add(paragraph.get("document_id"))

                if document_sheet is None:
                    result[paragraph.get("document_id")] = [
                        [paragraph.get("title"), paragraph.get("content"), "\n".join(problem_list)]
                    ]
                else:
                    document_sheet.append([paragraph.get("title"), paragraph.get("content"), "\n".join(problem_list)])
            result_document_dict = {}
            for d_name in document_dict:
                for index, d_id in enumerate(document_dict.get(d_name)):
                    result_document_dict[d_id] = d_name if index == 0 else d_name + str(index)
            return result, result_document_dict

    class ImportKnowledge(serializers.Serializer):
        user_id = serializers.UUIDField(required=True, label=_("user id"))
        workspace_id = serializers.CharField(required=True, label=_("workspace id"))
        folder_id = serializers.CharField(required=True, label=_("folder id"))
        scope = serializers.ChoiceField(
            required=False, label=_("scope"), default=KnowledgeScope.WORKSPACE, choices=KnowledgeScope.choices
        )

        @transaction.atomic
        def import_knowledge(self, file, is_import_tool=False, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
                KnowledgeImportRequest(data={"file": file}).is_valid(raise_exception=True)

            try:
                zf = zipfile.ZipFile(file)
            except zipfile.BadZipFile:
                raise AppApiException(500, _("Not a valid zip file"))

            namelist = zf.namelist()
            namelist_set = set(namelist)
            if "knowledge.json" not in namelist:
                raise AppApiException(500, _("Not a valid KB export file, missing knowledge.json"))
            if "knowledge.xlsx" not in namelist:
                raise AppApiException(500, _("Not a valid KB export file, missing knowledge.xlsx"))

            # knowledge.json -> knowledge
            knowledge_data = json.loads(zf.read("knowledge.json"))
            source_file_meta_map = {
                str(source_file.get("id")): source_file
                for source_file in knowledge_data.get("source_file_list", [])
                if source_file.get("id")
            }
            source_file_sheet_name_map = {
                source_file.get("sheet_name"): source_file
                for source_file in knowledge_data.get("source_file_list", [])
                if source_file.get("sheet_name")
            }
            workspace_id = self.data.get("workspace_id")
            user_id = self.data.get("user_id")
            knowledge_id = uuid.uuid7()
            folder_id = self.data.get("folder_id")
            knowledge = Knowledge(
                id=knowledge_id,
                name=knowledge_data.get("name", "Untitled"),
                desc=knowledge_data.get("desc", ""),
                type=knowledge_data.get("type", KnowledgeType.BASE),
                scope=self.data.get("scope", KnowledgeScope.WORKSPACE),
                meta=knowledge_data.get("meta", {}),
                file_size_limit=knowledge_data.get("file_size_limit", 100),
                file_count_limit=knowledge_data.get("file_count_limit", 50),
                embedding_model=None,
                user_id=user_id,
                workspace_id=workspace_id,
                folder_id=folder_id,
            )
            knowledge.save()

            # 图片
            old_to_new_file_map = {}
            for name in namelist:
                if name.startswith("oss/file/") and name != "oss/file/":
                    old_id = name.split("/")[-1]
                    if not old_id:
                        continue
                    file_bytes = zf.read(name)
                    new_file = File(
                        id=uuid.uuid7(),
                        file_name=old_id,
                        source_type=FileSourceType.KNOWLEDGE,
                        source_id=str(knowledge_id),
                        meta={},
                    )
                    new_file.save(bytea=file_bytes)
                    old_to_new_file_map[old_id] = str(new_file.id)

            # knowledge.xlsx -> doc + para + problem
            import openpyxl

            xlsx_bytes = io.BytesIO(zf.read("knowledge.xlsx"))
            workbook = openpyxl.load_workbook(xlsx_bytes)

            document_model_list = []
            paragraph_model_list = []
            problem_paragraph_object_list = []
            doc_tags_map = {}

            for sheet in workbook.worksheets:
                doc_name = sheet.title
                rows = list(sheet.iter_rows(min_row=2, values_only=True))
                if not rows:
                    continue

                # 首行文档元数据
                first_row = rows[0]
                tags_str = first_row[3] if len(first_row) > 3 and first_row[3] else ""
                hit_method = first_row[4] if len(first_row) > 4 and first_row[4] else "optimization"
                similarity = first_row[5] if len(first_row) > 5 and first_row[5] else 0.9
                doc_is_active = first_row[6] if len(first_row) > 6 and first_row[6] else "1"
                doc_type = (
                    first_row[8]
                    if len(first_row) > 8 and first_row[8]
                    else knowledge_data.get("type", KnowledgeType.BASE)
                )
                doc_meta_str = first_row[9] if len(first_row) > 9 and first_row[9] else "{}"

                try:
                    doc_meta = json.loads(doc_meta_str) if isinstance(doc_meta_str, str) else {}
                except (json.JSONDecodeError, TypeError):
                    doc_meta = {}

                char_length = sum(len(row[1] or "") for row in rows)
                document_id = uuid.uuid7()
                source_file_id = str(doc_meta["source_file_id"]) if doc_meta.get("source_file_id") else None
                source_file_meta = source_file_meta_map.get(source_file_id) if source_file_id else None
                if source_file_id is None and source_file_meta is None:
                    source_file_meta = source_file_sheet_name_map.get(doc_name)
                    source_file_id = str(source_file_meta.get("id")) if source_file_meta else None
                if source_file_id:
                    new_source_file_id = KnowledgeSerializer.Operate._restore_source_file(
                        zf,
                        namelist_set,
                        source_file_id,
                        source_file_meta,
                        document_id,
                    )
                    if new_source_file_id:
                        doc_meta["source_file_id"] = str(new_source_file_id)
                    else:
                        doc_meta.pop("source_file_id", None)
                document = Document(
                    id=document_id,
                    knowledge_id=knowledge_id,
                    name=doc_name,
                    char_length=char_length,
                    is_active=str(doc_is_active) == "1",
                    type=doc_type,
                    hit_handling_method=hit_method,
                    directly_return_similarity=float(similarity) if similarity else 0.9,
                    meta=doc_meta,
                )

                document_model_list.append(document)
                if tags_str:
                    doc_tags_map[document_id] = tags_str
                # 逐行创建 para + problem
                for row_idx, row in enumerate(rows):
                    title = str(row[0]) if len(row) > 0 and row[0] is not None else ""
                    content = str(row[1]) if len(row) > 1 and row[1] is not None else ""
                    problems_str = str(row[2]) if len(row) > 2 and row[2] is not None else ""
                    para_is_active = row[7] if len(row) > 7 and row[7] else "1"

                    # 图片 link 替换
                    for old_id, new_id in old_to_new_file_map.items():
                        content = content.replace(old_id, new_id)

                    if title.startswith("\ufeff"):
                        title = title[1:]
                    if content.startswith("\ufeff"):
                        content = content[1:]

                    paragraph_id = uuid.uuid7()
                    paragraph = Paragraph(
                        id=paragraph_id,
                        document_id=document_id,
                        knowledge_id=knowledge_id,
                        title=title,
                        content=content,
                        is_active=str(para_is_active) == "1",
                        position=row_idx + 1,
                        chunks=text_to_chunk(content),
                    )
                    paragraph_model_list.append(paragraph)

                    if problems_str:
                        if problems_str.startswith("\ufeff"):
                            problems_str = problems_str[1:]
                        for problem_content in problems_str.split("\n"):
                            problem_content = problem_content.strip()
                            if problem_content:
                                problem_paragraph_object_list.append(
                                    ProblemParagraphObject(knowledge_id, document_id, paragraph_id, problem_content)
                                )
            # bulk create
            QuerySet(Document).bulk_create(document_model_list) if len(document_model_list) > 0 else None
            QuerySet(Paragraph).bulk_create(paragraph_model_list) if len(paragraph_model_list) > 0 else None

            # 问题
            problem_model_list, problem_paragraph_mapping_list = ProblemParagraphManage(
                problem_paragraph_object_list, knowledge_id
            ).to_problem_model_list()
            bulk_create_in_batches(Problem, problem_model_list, batch_size=1000)
            bulk_create_in_batches(ProblemParagraphMapping, problem_paragraph_mapping_list, batch_size=1000)

            # Tag
            tag_list = knowledge_data.get("tags", [])
            if tag_list:
                tag_model_list = []
                tag_key_value_to_model = {}
                for tag in tag_list:
                    tag_model = Tag(id=uuid.uuid7(), knowledge_id=knowledge_id, key=tag["key"], value=tag["value"])
                    tag_model_list.append(tag_model)

                    tag_key_value_to_model[f"{tag['key']}:{tag['value']}"] = tag_model
                QuerySet(Tag).bulk_create(tag_model_list)

                # Document_Tag
                document_tag_model_list = []
                for doc_id, tags_str in doc_tags_map.items():
                    for tag_str in tags_str.split("|"):
                        tag_str = tag_str.strip()
                        if tag_str and tag_str in tag_key_value_to_model:
                            document_tag_model_list.append(
                                DocumentTag(
                                    id=uuid.uuid7(), document_id=doc_id, tag_id=tag_key_value_to_model[tag_str].id
                                )
                            )
                QuerySet(DocumentTag).bulk_create(document_tag_model_list) if len(document_tag_model_list) > 0 else None

            # Termbase
            terms = knowledge_data.get("termbase", [])
            if terms:
                termbase_instance_list = [
                    Termbase(id=uuid.uuid7(), knowledge_id=knowledge_id, content=content) for content in terms
                ]
                QuerySet(Termbase).bulk_create(termbase_instance_list) if len(termbase_instance_list) > 0 else None

            # 授权 + 资源映射
            UserResourcePermissionSerializer(
                data={
                    "workspace_id": self.data.get("workspace_id"),
                    "user_id": self.data.get("user_id"),
                    "auth_target_type": AuthTargetType.KNOWLEDGE.value,
                }
            ).auth_resource(str(knowledge_id))

            update_resource_mapping_by_knowledge(str(knowledge_id))

            zf.close()
            return {"knowledge_id": str(knowledge_id), "type": knowledge.type}

    class Create(serializers.Serializer):
        user_id = serializers.UUIDField(required=True, label=_("user id"))
        workspace_id = serializers.CharField(required=True, label=_("workspace id"))
        scope = serializers.ChoiceField(
            required=False, label=_("scope"), default=KnowledgeScope.WORKSPACE, choices=KnowledgeScope.choices
        )

        @staticmethod
        def post_embedding_knowledge(document_list, knowledge_id):
            model_id = get_embedding_model_id_by_knowledge_id(knowledge_id)
            embedding_by_knowledge.delay(knowledge_id, model_id)
            return document_list

        @post(post_function=post_embedding_knowledge)
        @transaction.atomic
        def save_base(self, instance, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
                KnowledgeBaseCreateRequest(data=instance).is_valid(raise_exception=True)
            folder_id = instance.get("folder_id", self.data.get("workspace_id"))

            knowledge_id = uuid.uuid7()
            knowledge = Knowledge(
                id=knowledge_id,
                name=instance.get("name"),
                workspace_id=self.data.get("workspace_id"),
                desc=instance.get("desc"),
                type=instance.get("type", KnowledgeType.BASE),
                user_id=self.data.get("user_id"),
                scope=self.data.get("scope", KnowledgeScope.WORKSPACE),
                folder_id=folder_id,
                embedding_model_id=instance.get("embedding_model_id"),
                meta=instance.get("meta", {}),
            )

            document_model_list = []
            paragraph_model_list = []
            problem_paragraph_object_list = []
            # 插入文档
            for document in instance.get("documents") if "documents" in instance else []:
                document_paragraph_dict_model = DocumentSerializers.Create.get_document_paragraph_model(
                    knowledge_id, self.data.get("user_id"), document
                )
                document_model_list.append(document_paragraph_dict_model.get("document"))
                for paragraph in document_paragraph_dict_model.get("paragraph_model_list"):
                    paragraph_model_list.append(paragraph)
                for problem_paragraph_object in document_paragraph_dict_model.get("problem_paragraph_object_list"):
                    problem_paragraph_object_list.append(problem_paragraph_object)

            problem_model_list, problem_paragraph_mapping_list = ProblemParagraphManage(
                problem_paragraph_object_list, knowledge_id
            ).to_problem_model_list()
            # 插入知识库
            knowledge.save()
            # 插入文档
            QuerySet(Document).bulk_create(document_model_list) if len(document_model_list) > 0 else None
            # 批量插入段落
            QuerySet(Paragraph).bulk_create(paragraph_model_list) if len(paragraph_model_list) > 0 else None
            # 批量插入问题
            QuerySet(Problem).bulk_create(problem_model_list) if len(problem_model_list) > 0 else None
            # 批量插入关联问题
            QuerySet(ProblemParagraphMapping).bulk_create(problem_paragraph_mapping_list) if len(
                problem_paragraph_mapping_list
            ) > 0 else None
            # 自动资源给授权当前用户
            UserResourcePermissionSerializer(
                data={
                    "workspace_id": self.data.get("workspace_id"),
                    "user_id": self.data.get("user_id"),
                    "auth_target_type": AuthTargetType.KNOWLEDGE.value,
                }
            ).auth_resource(str(knowledge_id))
            update_resource_mapping_by_knowledge(str(knowledge_id))
            return {
                **KnowledgeModelSerializer(knowledge).data,
                "user_id": self.data.get("user_id"),
                "document_list": document_model_list,
                "document_count": len(document_model_list),
                "char_length": reduce(lambda x, y: x + y, [d.char_length for d in document_model_list], 0),
            }, knowledge_id

    class HitTest(serializers.Serializer):
        workspace_id = serializers.CharField(required=True, label=_("workspace id"))
        knowledge_id = serializers.UUIDField(required=True, label=_("id"))
        user_id = serializers.UUIDField(required=False, label=_("user id"))
        query_text = serializers.CharField(required=True, label=_("query text"))
        top_number = serializers.IntegerField(required=True, max_value=10000, min_value=1, label=_("top number"))
        similarity = serializers.FloatField(required=True, max_value=2, min_value=0, label=_("similarity"))
        search_mode = serializers.CharField(
            required=True,
            label=_("search mode"),
            validators=[
                validators.RegexValidator(
                    regex=re.compile("^embedding|keywords|blend$"),
                    message=_("The type only supports embedding|keywords|blend"),
                    code=500,
                )
            ],
        )

        def is_valid(self, *, raise_exception=True):
            super().is_valid(raise_exception=True)
            workspace_id = self.data.get("workspace_id")
            query_set = QuerySet(Knowledge).filter(id=self.data.get("knowledge_id"))
            if workspace_id:
                query_set = query_set.filter(workspace_id=workspace_id)
            if not query_set.exists():
                raise AppApiException(500, _("Knowledge id does not exist"))
            if not QuerySet(Knowledge).filter(id=self.data.get("knowledge_id")).exists():
                raise AppApiException(300, _("id does not exist"))

        def hit_test(self):
            self.is_valid()
            vector = VectorStore.get_embedding_vector()
            exclude_document_id_list = [
                str(document.id)
                for document in QuerySet(Document).filter(knowledge_id=self.data.get("knowledge_id"), is_active=False)
            ]
            model = get_embedding_model_by_knowledge_id(self.data.get("knowledge_id"))
            # 向量库检索
            hit_list = vector.hit_test(
                self.data.get("query_text"),
                [self.data.get("knowledge_id")],
                exclude_document_id_list,
                self.data.get("top_number"),
                self.data.get("similarity"),
                SearchMode(self.data.get("search_mode")),
                model,
            )
            hit_dict = reduce(lambda x, y: {**x, **y}, [{hit.get("paragraph_id"): hit} for hit in hit_list], {})
            p_list = list_paragraph([h.get("paragraph_id") for h in hit_list])
            return [
                {
                    **p,
                    "similarity": hit_dict.get(p.get("id")).get("similarity"),
                    "comprehensive_score": hit_dict.get(p.get("id")).get("comprehensive_score"),
                }
                for p in p_list
            ]

    class BatchHitTest(serializers.Serializer):
        workspace_id = serializers.CharField(required=True, label=_("workspace id"))
        knowledge_id_list = serializers.ListField(
            required=True,
            allow_empty=False,
            child=serializers.UUIDField(required=True, label=_("knowledge id")),
            label=_("knowledge id list"),
        )
        user_id = serializers.UUIDField(required=False, label=_("user id"))
        query_text = serializers.CharField(required=True, label=_("query text"))
        top_number = serializers.IntegerField(required=True, max_value=10000, min_value=1, label=_("top number"))
        similarity = serializers.FloatField(required=True, max_value=2, min_value=0, label=_("similarity"))
        search_mode = serializers.CharField(
            required=True,
            label=_("search mode"),
            validators=[
                validators.RegexValidator(
                    regex=re.compile("^embedding|keywords|blend$"),
                    message=_("The type only supports embedding|keywords|blend"),
                    code=500,
                )
            ],
        )

        def is_valid(self, *, raise_exception=True):
            super().is_valid(raise_exception=True)
            workspace_id = self.validated_data.get("workspace_id")
            knowledge_id_list = [str(knowledge_id) for knowledge_id in self.validated_data.get("knowledge_id_list")]
            knowledge_set = QuerySet(Knowledge).filter(id__in=knowledge_id_list)
            if workspace_id:
                knowledge_set = knowledge_set.filter(workspace_id=workspace_id)
            existing_id_set = {str(knowledge.id) for knowledge in knowledge_set}
            if len(existing_id_set) != len(set(knowledge_id_list)):
                raise AppApiException(500, _("Knowledge id does not exist"))

        def hit_test(self):
            self.is_valid()
            knowledge_id_list = [str(knowledge_id) for knowledge_id in self.validated_data.get("knowledge_id_list")]
            vector = VectorStore.get_embedding_vector()
            exclude_document_id_list = [
                str(document.id)
                for document in QuerySet(Document).filter(knowledge_id__in=knowledge_id_list, is_active=False)
            ]
            try:
                model = get_embedding_model_by_knowledge_id_list(knowledge_id_list)
            except Exception as e:
                raise AppApiException(500, str(e)) from e
            hit_list = vector.hit_test(
                self.validated_data.get("query_text"),
                knowledge_id_list,
                exclude_document_id_list,
                self.validated_data.get("top_number"),
                self.validated_data.get("similarity"),
                SearchMode(self.validated_data.get("search_mode")),
                model,
            )
            hit_dict = reduce(lambda x, y: {**x, **y}, [{hit.get("paragraph_id"): hit} for hit in hit_list], {})
            p_list = list_paragraph([h.get("paragraph_id") for h in hit_list])
            return [
                {
                    **p,
                    "similarity": hit_dict.get(p.get("id")).get("similarity"),
                    "comprehensive_score": hit_dict.get(p.get("id")).get("comprehensive_score"),
                }
                for p in p_list
            ]

    class Tags(serializers.Serializer):
        workspace_id = serializers.CharField(required=True, label=_("workspace id"))
        user_id = serializers.UUIDField(required=True, label=_("user id"))
        knowledge_ids = serializers.ListField(
            required=True, label=_("knowledge ids"), child=serializers.UUIDField(required=True, label=_("id"))
        )

        def list(self):
            self.is_valid(raise_exception=True)
            if self.data.get("name"):
                name = self.data.get("name")
                tags = (
                    QuerySet(Tag)
                    .filter(knowledge_id__in=self.data.get("knowledge_ids"))
                    .filter(Q(key__icontains=name) | Q(value__icontains=name))
                    .values("key", "value", "id", "create_time", "update_time")
                    .order_by("create_time", "key", "value")
                )
            else:
                # 获取所有标签，按创建时间排序保持稳定顺序
                tags = (
                    QuerySet(Tag)
                    .filter(knowledge_id__in=self.data.get("knowledge_ids"))
                    .values("key", "value", "id", "create_time", "update_time")
                    .order_by("create_time", "key", "value")
                )

            # 按key分组
            grouped_tags = defaultdict(list)
            for tag in tags:
                grouped_tags[tag["key"]].append(
                    {
                        "id": tag["id"],
                        "value": tag["value"],
                        "create_time": tag["create_time"],
                        "update_time": tag["update_time"],
                    }
                )

            # 转换为期望的格式，保持key的顺序
            result = []
            # 按key排序以确保结果顺序一致
            for key in sorted(grouped_tags.keys()):
                values = grouped_tags[key]
                # 按创建时间对values进行排序
                values.sort(key=lambda x: x["create_time"])
                result.append(
                    {
                        "key": key,
                        "values": values,
                    }
                )

            return result


class KnowledgeBatchOperateSerializer(serializers.Serializer):
    workspace_id = serializers.CharField(required=True, label=_("workspace id"))

    def is_valid(self, *, raise_exception=False):
        super().is_valid(raise_exception=True)

    @transaction.atomic
    def batch_delete(self, instance: Dict, with_valid=True):
        if with_valid:
            BatchSerializer(data=instance).is_valid(model=Knowledge, raise_exception=True)
            self.is_valid(raise_exception=True)
        id_list = instance.get("id_list")
        workspace_id = self.data.get("workspace_id")
        knowledge_query_set = QuerySet(Knowledge).filter(id__in=id_list, workspace_id=workspace_id)

        # 删除所有关联
        QuerySet(Document).filter(knowledge__in=knowledge_query_set).delete()
        QuerySet(ProblemParagraphMapping).filter(knowledge__in=knowledge_query_set).delete()
        QuerySet(Paragraph).filter(knowledge__in=knowledge_query_set).delete()
        QuerySet(Problem).filter(knowledge__in=knowledge_query_set).delete()
        QuerySet(WorkspaceUserResourcePermission).filter(target__in=id_list).delete()

        for k_id in id_list:
            drop_knowledge_index(knowledge_id=k_id)
            delete_embedding_by_knowledge(k_id)

        File.objects.filter(source_id__in=id_list).delete()
        QuerySet(ResourceMapping).filter(Q(target_id__in=id_list) | Q(source_id__in=id_list)).delete()

        knowledge_query_set.delete()
        return True

    def batch_move(self, instance: Dict, with_valid=True):
        if with_valid:
            BatchMoveSerializer(data=instance).is_valid(model=Knowledge, raise_exception=True)
            self.is_valid(raise_exception=True)
        id_list = instance.get("id_list")
        folder_id = instance.get("folder_id")
        workspace_id = self.data.get("workspace_id")

        QuerySet(Knowledge).filter(id__in=id_list, workspace_id=workspace_id).update(folder_id=folder_id)
        return True
